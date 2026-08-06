from __future__ import annotations

import asyncio
import csv
import html as html_lib
import io
import json
import time
from collections import defaultdict, deque
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import urljoin
from xml.etree import ElementTree

import httpx
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import Settings
from app.services.extractor import ExtractedPage, extract_page
from app.services.normalizer import canonical_url, domain_of, is_same_or_subdomain, root_url, url_is_allowed
from app.services.robots import RobotsCache

_DOCUMENT_EXTENSIONS = (".pdf", ".csv", ".tsv", ".xlsx", ".json", ".xml")
_CONTACT_PATHS = (
    "contact",
    "contacts",
    "contact-us",
    "contatti",
    "contatto",
    "kontakt",
    "kontakte",
    "contacto",
    "contactos",
    "contato",
    "impressum",
    "legal-notice",
    "about",
    "chi-siamo",
    "team",
    "staff",
    "locations",
    "sedi",
)
_CONTACT_HINTS = (
    "contact",
    "contatt",
    "kontakt",
    "contacto",
    "contato",
    "impressum",
    "legal",
    "about",
    "chi-siamo",
    "team",
    "staff",
    "location",
    "sedi",
    "clinic",
    "office",
    "email",
    "albo",
    "elenco",
    "directory",
    "dataset",
)


@dataclass
class PageResult:
    url: str
    status_code: int | None
    html: str | None
    action: str
    detail: str | None = None


@dataclass
class OrganizationCandidate:
    organization: str | None
    website: str
    domain: str
    source_url: str
    emails: list[str] = field(default_factory=list)
    phones: list[str] = field(default_factory=list)
    whatsapp: list[str] = field(default_factory=list)
    address: str | None = None
    city: str | None = None
    specialties: list[str] = field(default_factory=list)
    pages_crawled: int = 0
    logs: list[PageResult] = field(default_factory=list)


class DomainRateLimiter:
    def __init__(self, default_delay: float) -> None:
        self.default_delay = default_delay
        self._locks: defaultdict[str, asyncio.Lock] = defaultdict(asyncio.Lock)
        self._last_request: dict[str, float] = {}

    async def wait(self, domain: str, requested_delay: float | None = None) -> None:
        delay = max(self.default_delay, requested_delay or 0)
        async with self._locks[domain]:
            elapsed = time.monotonic() - self._last_request.get(domain, 0.0)
            if elapsed < delay:
                await asyncio.sleep(delay - elapsed)
            self._last_request[domain] = time.monotonic()


class ContactCrawler:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        headers = {
            "User-Agent": settings.crawler_user_agent,
            "Accept": (
                "text/html,application/xhtml+xml,application/pdf,text/csv,application/json,"
                "application/xml,text/xml,application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet;q=0.9,*/*;q=0.5"
            ),
            "Accept-Language": "en,it;q=0.8,*;q=0.5",
        }
        self.client = httpx.AsyncClient(
            headers=headers,
            timeout=settings.crawler_timeout_seconds,
            follow_redirects=False,
        )
        self.robots = RobotsCache(self.client, settings.crawler_user_agent, settings.respect_robots_txt)
        self.limiter = DomainRateLimiter(settings.crawler_delay_seconds)

    async def close(self) -> None:
        await self.client.aclose()

    async def _safe(self, url: str) -> tuple[bool, str | None]:
        return await asyncio.to_thread(url_is_allowed, url, self.settings.blocked_domain_set)

    async def _request(self, url: str, *, enforce_robots: bool = True) -> tuple[str, httpx.Response | None, str | None]:
        current = url
        for _ in range(6):
            allowed, reason = await self._safe(current)
            if not allowed:
                return current, None, reason or "URL bloccata"
            requested_delay: float | None = None
            if enforce_robots:
                robots = await self.robots.can_fetch(current)
                if not robots.allowed:
                    return current, None, "robots.txt non consente la scansione"
                requested_delay = robots.crawl_delay
            await self.limiter.wait(domain_of(current), requested_delay)
            try:
                response = await self.client.get(current)
            except httpx.HTTPError as exc:
                return current, None, str(exc)
            if response.status_code in {301, 302, 303, 307, 308}:
                location = response.headers.get("location")
                if not location:
                    return current, response, "redirect senza Location"
                redirected = canonical_url(location, current)
                if not redirected:
                    return current, response, "redirect non valido"
                current = redirected
                continue
            return current, response, None
        return current, None, "troppi redirect"

    @staticmethod
    def _html_pre(text: str) -> str:
        return f"<html><body><pre>{html_lib.escape(text)}</pre></body></html>"

    def _pdf_to_html(self, content: bytes) -> str | None:
        if not self.settings.pdf_enabled:
            return None
        try:
            reader = PdfReader(io.BytesIO(content), strict=False)
            texts: list[str] = []
            for page in reader.pages[: self.settings.pdf_max_pages]:
                value = page.extract_text() or ""
                if value.strip():
                    texts.append(value)
            text = "\n".join(texts).strip()
            return self._html_pre(text) if text else None
        except Exception:
            return None

    @staticmethod
    def _decode_text(content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")

    def _csv_to_html(self, content: bytes, delimiter: str | None = None) -> str | None:
        if not self.settings.structured_documents_enabled:
            return None
        text = self._decode_text(content)
        sample = text[:20_000]
        if delimiter is None:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                delimiter = ";"
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        lines: list[str] = []
        for index, row in enumerate(reader):
            if index >= self.settings.structured_document_max_rows:
                break
            values = [" ".join(str(value).split()) for value in row if str(value).strip()]
            if values:
                lines.append(" | ".join(values))
        flattened = "\n".join(lines).strip()
        return self._html_pre(flattened) if flattened else None

    def _xlsx_to_html(self, content: bytes) -> str | None:
        if not self.settings.structured_documents_enabled:
            return None
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            lines: list[str] = []
            rows_seen = 0
            for sheet in workbook.worksheets:
                lines.append(f"FOGLIO: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    if rows_seen >= self.settings.structured_document_max_rows:
                        break
                    values = [" ".join(str(value).split()) for value in row if value not in (None, "")]
                    if values:
                        lines.append(" | ".join(values))
                    rows_seen += 1
                if rows_seen >= self.settings.structured_document_max_rows:
                    break
            workbook.close()
            flattened = "\n".join(lines).strip()
            return self._html_pre(flattened) if flattened else None
        except Exception:
            return None

    def _json_to_html(self, content: bytes) -> str | None:
        if not self.settings.structured_documents_enabled:
            return None
        try:
            payload = json.loads(self._decode_text(content))
            text = json.dumps(payload, ensure_ascii=False, indent=1)
            return self._html_pre(text[:8_000_000])
        except (json.JSONDecodeError, TypeError):
            return None

    def _xml_to_html(self, content: bytes) -> str | None:
        if not self.settings.structured_documents_enabled:
            return None
        try:
            root = ElementTree.fromstring(content)
            text = "\n".join(value.strip() for value in root.itertext() if value.strip())
            return self._html_pre(text[:8_000_000]) if text else None
        except ElementTree.ParseError:
            return None

    async def _fetch_http(self, url: str) -> PageResult:
        current, response, error = await self._request(url, enforce_robots=True)
        if response is None:
            action = "robots_blocked" if error and "robots.txt" in error else "error"
            if error and ("blocc" in error or "non valido" in error):
                action = "blocked"
            return PageResult(current, None, None, action, error)
        if error:
            return PageResult(current, response.status_code, None, "error", error)
        if response.status_code >= 400:
            return PageResult(current, response.status_code, None, "error", f"HTTP {response.status_code}")

        content = response.content
        if len(content) > self.settings.crawler_max_response_bytes:
            return PageResult(current, response.status_code, None, "ignored", "risorsa oltre il limite di dimensione")
        content_type = response.headers.get("content-type", "").lower()
        lowered_url = current.casefold().split("?", 1)[0]

        if "application/pdf" in content_type or lowered_url.endswith(".pdf"):
            html = await asyncio.to_thread(self._pdf_to_html, content)
            if html:
                return PageResult(current, response.status_code, html, "pdf", "PDF pubblico estratto")
            return PageResult(current, response.status_code, None, "ignored", "PDF senza testo estraibile")
        if "text/csv" in content_type or lowered_url.endswith(".csv"):
            html = await asyncio.to_thread(self._csv_to_html, content, None)
            return PageResult(current, response.status_code, html, "csv", "CSV pubblico estratto")
        if "tab-separated-values" in content_type or lowered_url.endswith(".tsv"):
            html = await asyncio.to_thread(self._csv_to_html, content, "\t")
            return PageResult(current, response.status_code, html, "csv", "TSV pubblico estratto")
        if (
            "spreadsheetml.sheet" in content_type
            or lowered_url.endswith(".xlsx")
        ):
            html = await asyncio.to_thread(self._xlsx_to_html, content)
            return PageResult(current, response.status_code, html, "spreadsheet", "XLSX pubblico estratto")
        if "application/json" in content_type or lowered_url.endswith(".json"):
            html = await asyncio.to_thread(self._json_to_html, content)
            return PageResult(current, response.status_code, html, "json", "JSON pubblico estratto")
        if (
            "application/xml" in content_type
            or "text/xml" in content_type
            or lowered_url.endswith(".xml")
        ):
            html = await asyncio.to_thread(self._xml_to_html, content)
            return PageResult(current, response.status_code, html, "xml", "XML pubblico estratto")
        if (
            "text/html" not in content_type
            and "application/xhtml+xml" not in content_type
            and "text/plain" not in content_type
        ):
            return PageResult(current, response.status_code, None, "ignored", f"content-type {content_type or 'sconosciuto'}")
        encoding = response.encoding or "utf-8"
        try:
            html = content.decode(encoding, errors="replace")
        except LookupError:
            html = content.decode("utf-8", errors="replace")
        if "text/plain" in content_type:
            html = self._html_pre(html)
        return PageResult(current, response.status_code, html, "fetched")

    async def _fetch_playwright(self, url: str) -> PageResult:
        if not self.settings.playwright_enabled:
            return PageResult(url, None, None, "ignored", "fallback Playwright disabilitato")
        try:
            from playwright.async_api import async_playwright
        except ImportError:
            return PageResult(url, None, None, "error", "Playwright non installato")
        allowed, reason = await self._safe(url)
        if not allowed:
            return PageResult(url, None, None, "blocked", reason)
        robots = await self.robots.can_fetch(url)
        if not robots.allowed:
            return PageResult(url, None, None, "robots_blocked", "robots.txt non consente la scansione")
        await self.limiter.wait(domain_of(url), robots.crawl_delay)
        try:
            async with async_playwright() as playwright:
                browser = await playwright.chromium.launch(headless=True)
                context = await browser.new_context(user_agent=self.settings.crawler_user_agent)

                async def guard_request(route):
                    request_url = route.request.url
                    if request_url.startswith(("data:", "blob:")):
                        await route.continue_()
                        return
                    request_allowed, _ = await self._safe(request_url)
                    if request_allowed:
                        await route.continue_()
                    else:
                        await route.abort()

                await context.route("**/*", guard_request)
                page = await context.new_page()
                response = await page.goto(
                    url,
                    wait_until="domcontentloaded",
                    timeout=self.settings.crawler_timeout_seconds * 1000,
                )
                await page.wait_for_timeout(700)
                final_url = canonical_url(page.url)
                if not final_url:
                    await browser.close()
                    return PageResult(url, None, None, "error", "URL finale non valida")
                final_allowed, final_reason = await self._safe(final_url)
                if not final_allowed:
                    await browser.close()
                    return PageResult(final_url, None, None, "blocked", final_reason)
                html = await page.content()
                status_code = response.status if response else None
                await browser.close()
                if len(html.encode("utf-8")) > self.settings.crawler_max_response_bytes:
                    return PageResult(final_url, status_code, None, "ignored", "pagina oltre il limite di dimensione")
                return PageResult(final_url, status_code, html, "playwright")
        except Exception as exc:
            return PageResult(url, None, None, "error", f"Playwright: {exc}")

    async def fetch(self, url: str) -> PageResult:
        result = await self._fetch_http(url)
        if result.action in {"pdf", "csv", "spreadsheet", "json", "xml"}:
            return result
        if result.html and len(result.html) >= 1_000:
            return result
        if result.action in {"blocked", "robots_blocked", "ignored"}:
            return result
        fallback = await self._fetch_playwright(result.url)
        return fallback if fallback.html else result

    @staticmethod
    def _sitemap_score(url: str) -> tuple[int, int, str]:
        lowered = url.casefold()
        score = sum(20 for hint in _CONTACT_HINTS if hint in lowered)
        if lowered.split("?", 1)[0].endswith(_DOCUMENT_EXTENSIONS):
            score += 12
        return (-score, lowered.count("/"), url)

    async def _sitemap_urls(self, base_root: str, parent_domain: str) -> list[str]:
        if not self.settings.sitemap_enabled or self.settings.sitemap_max_urls_per_domain <= 0:
            return []
        candidates = [urljoin(f"{base_root}/", "sitemap.xml"), urljoin(f"{base_root}/", "sitemap_index.xml")]
        robots_url = urljoin(f"{base_root}/", "robots.txt")
        _, robots_response, _ = await self._request(robots_url, enforce_robots=False)
        if robots_response and robots_response.status_code < 400:
            for line in robots_response.text.splitlines():
                if line.casefold().startswith("sitemap:"):
                    value = canonical_url(line.split(":", 1)[1].strip(), base_root)
                    if value:
                        candidates.append(value)

        discovered: list[str] = []
        sitemap_queue: deque[str] = deque(dict.fromkeys(candidates))
        visited_sitemaps: set[str] = set()
        while sitemap_queue and len(visited_sitemaps) < 8:
            sitemap_url = sitemap_queue.popleft()
            if sitemap_url in visited_sitemaps:
                continue
            visited_sitemaps.add(sitemap_url)
            _, response, _ = await self._request(sitemap_url, enforce_robots=False)
            if not response or response.status_code >= 400 or len(response.content) > 8_000_000:
                continue
            try:
                root = ElementTree.fromstring(response.content)
            except ElementTree.ParseError:
                continue
            for element in root.iter():
                if not element.tag.casefold().endswith("loc") or not element.text:
                    continue
                target = canonical_url(element.text.strip(), base_root)
                if not target:
                    continue
                host = domain_of(target).removeprefix("www.")
                if not (
                    is_same_or_subdomain(host, parent_domain)
                    or is_same_or_subdomain(parent_domain, host)
                ):
                    continue
                path = target.casefold().split("?", 1)[0]
                if path.endswith((".xml.gz",)) or (
                    path.endswith(".xml") and "sitemap" in path
                ):
                    sitemap_queue.append(target)
                    continue
                if any(hint in target.casefold() for hint in _CONTACT_HINTS) or path.endswith(_DOCUMENT_EXTENSIONS):
                    discovered.append(target)

        unique = list(dict.fromkeys(discovered))
        unique.sort(key=self._sitemap_score)
        return unique[: self.settings.sitemap_max_urls_per_domain]

    async def crawl_domain(
        self,
        start_url: str,
        country: str | None = None,
        keywords: list[str] | None = None,
    ) -> OrganizationCandidate:
        normalized_start = canonical_url(start_url)
        if not normalized_start:
            raise ValueError("URL iniziale non valida")
        parent_domain = domain_of(normalized_start).removeprefix("www.")
        base_root = root_url(normalized_start)
        sitemap_urls = await self._sitemap_urls(base_root, parent_domain)
        seeds = [normalized_start, base_root]
        seeds.extend(urljoin(f"{base_root}/", path) for path in _CONTACT_PATHS)
        seeds.extend(sitemap_urls)
        queue: deque[str] = deque(dict.fromkeys(seeds))
        visited: set[str] = set()
        pages: list[tuple[str, ExtractedPage]] = []
        logs: list[PageResult] = []
        if sitemap_urls:
            logs.append(PageResult(base_root, None, None, "sitemap", f"{len(sitemap_urls)} URL utili dalla sitemap"))
        attempt_limit = max(self.settings.crawler_max_pages_per_domain * 7, 24)

        while (
            queue
            and len(pages) < self.settings.crawler_max_pages_per_domain
            and len(visited) < attempt_limit
        ):
            url = queue.popleft()
            canonical = canonical_url(url)
            if not canonical or canonical in visited:
                continue
            current_domain = domain_of(canonical).removeprefix("www.")
            if not is_same_or_subdomain(current_domain, parent_domain) and not is_same_or_subdomain(parent_domain, current_domain):
                continue
            visited.add(canonical)
            result = await self.fetch(canonical)
            logs.append(result)
            if not result.html:
                continue
            extracted = extract_page(result.html, result.url, country=country, keywords=keywords)
            pages.append((result.url, extracted))

            for href in extracted.relevant_links:
                target = canonical_url(href, result.url)
                if not target or target in visited:
                    continue
                target_domain = domain_of(target).removeprefix("www.")
                if is_same_or_subdomain(target_domain, parent_domain) or is_same_or_subdomain(parent_domain, target_domain):
                    queue.append(target)

            if self.settings.keep_page_snapshots:
                safe_name = parent_domain.replace(".", "_") + f"_{len(pages)}.html"
                path = Path(self.settings.snapshot_directory) / safe_name
                path.write_text(result.html, encoding="utf-8")

        organization = next((page.organization for _, page in pages if page.organization), None)
        emails = sorted({value for _, page in pages for value in page.emails})
        phones = sorted({value for _, page in pages for value in page.phones})
        whatsapp = sorted({value for _, page in pages for value in page.whatsapp})
        address = next((page.address for _, page in pages if page.address), None)
        city = next((page.city for _, page in pages if page.city), None)
        specialties = sorted({value for _, page in pages for value in page.specialties})
        source_url = next((url for url, page in pages if page.emails or page.phones), normalized_start)

        return OrganizationCandidate(
            organization=organization,
            website=root_url(source_url),
            domain=parent_domain,
            source_url=source_url,
            emails=emails,
            phones=phones,
            whatsapp=whatsapp,
            address=address,
            city=city,
            specialties=specialties,
            pages_crawled=len(pages),
            logs=logs,
        )
