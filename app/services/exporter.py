from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models import Contact

HEADERS = [
    "ID", "Struttura", "Categoria", "Paese", "Città", "Indirizzo", "Sito", "Dominio",
    "Email principale", "Email secondarie", "Telefoni", "WhatsApp", "Specializzazioni",
    "URL fonte", "Tipo fonte", "Affidabilità", "Stato", "Verificato il", "Note",
]


def _rows(db: Session, job_id: str | None = None) -> list[list[str | int]]:
    statement = select(Contact).order_by(Contact.country, Contact.city, Contact.organization)
    if job_id:
        statement = statement.where(Contact.job_id == job_id)
    contacts = list(db.scalars(statement))
    result: list[list[str | int]] = []
    for item in contacts:
        result.append([
            item.id,
            item.organization,
            item.category or "",
            item.country or "",
            item.city or "",
            item.address or "",
            item.website,
            item.domain,
            item.emails[0] if item.emails else "",
            "; ".join(item.emails[1:]),
            "; ".join(item.phones),
            "; ".join(item.whatsapp),
            "; ".join(item.specialties),
            item.source_url,
            item.source_type,
            item.reliability,
            item.status,
            item.verified_at.isoformat(),
            item.notes or "",
        ])
    return result


def export_csv(db: Session, job_id: str | None = None) -> Path:
    settings = get_settings()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{job_id}" if job_id else ""
    path = settings.export_directory / f"contatti{suffix}_{stamp}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(_rows(db, job_id))
    return path


def export_xlsx(db: Session, job_id: str | None = None) -> Path:
    settings = get_settings()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{job_id}" if job_id else ""
    path = settings.export_directory / f"contatti{suffix}_{stamp}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contatti"
    sheet.append(HEADERS)
    for row in _rows(db, job_id):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [8, 32, 24, 18, 20, 35, 35, 24, 32, 36, 24, 24, 38, 42, 18, 14, 18, 24, 45]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)
    return path
