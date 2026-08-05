from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Contact
from app.services.normalizer import canonical_url, contact_fingerprint, domain_of, normalize_email, root_url


def _key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _split(value: object) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in re.split(r"[;,\n]+", str(value)) if part.strip()]


def import_xlsx(db: Session, path: Path) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Contatti verificati"] if "Contatti verificati" in workbook.sheetnames else workbook.active

    header_row = None
    headers: dict[str, int] = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        keys = [_key(cell) for cell in row]
        if any(key in {"email principale", "email", "e mail principale"} for key in keys):
            header_row = row_index
            headers = {key: index for index, key in enumerate(keys) if key}
            break
    if not header_row:
        raise ValueError("Intestazione non riconosciuta: manca la colonna Email principale")

    def value(row: tuple, *names: str):
        for name in names:
            index = headers.get(_key(name))
            if index is not None and index < len(row):
                return row[index]
        return None

    imported = 0
    duplicates = 0
    existing_fingerprints = set(db.scalars(select(Contact.fingerprint)))
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        primary = normalize_email(str(value(row, "Email principale", "Email") or ""))
        secondary = [normalize_email(item) for item in _split(value(row, "Email secondaria", "Email secondarie"))]
        emails = [item for item in [primary, *secondary] if item]
        website_raw = str(value(row, "Sito ufficiale", "Sito", "Website") or "").strip()
        source_raw = str(value(row, "URL fonte", "Fonte", "Source URL") or website_raw).strip()
        website = canonical_url(website_raw) or canonical_url(source_raw)
        if not website or not emails:
            continue
        website = root_url(website)
        domain = domain_of(website).removeprefix("www.")
        fingerprint = contact_fingerprint(domain, emails, _split(value(row, "Telefono", "Telefoni")))
        verified = value(row, "Verificato il", "Data verifica")
        verified_at = verified if isinstance(verified, datetime) else datetime.now(timezone.utc)
        reliability_raw = str(value(row, "Affidabilità") or "medium").lower()
        reliability = "high" if reliability_raw in {"alta", "high"} else ("low" if reliability_raw in {"bassa", "low"} else "medium")
        if fingerprint in existing_fingerprints:
            duplicates += 1
            continue
        contact = Contact(
            organization=str(value(row, "Struttura", "Organizzazione", "Nome") or domain)[:300],
            category=str(value(row, "Categoria") or "")[:200] or None,
            country=str(value(row, "Paese", "Country") or "")[:100] or None,
            city=str(value(row, "Città / area", "Città", "City") or "")[:150] or None,
            address=str(value(row, "Indirizzo", "Address") or "") or None,
            website=website,
            domain=domain,
            emails=list(dict.fromkeys(emails)),
            phones=_split(value(row, "Telefono", "Telefoni")),
            whatsapp=_split(value(row, "WhatsApp")),
            specialties=_split(value(row, "Specializzazioni")),
            source_url=canonical_url(source_raw) or website,
            source_type=str(value(row, "Tipo fonte") or "official_website")[:80],
            reliability=reliability,
            status="verified_public",
            notes=str(value(row, "Note") or "Importato da Excel") or None,
            fingerprint=fingerprint,
            verified_at=verified_at,
        )
        db.add(contact)
        existing_fingerprints.add(fingerprint)
        imported += 1
    db.commit()
    workbook.close()
    return imported, duplicates
